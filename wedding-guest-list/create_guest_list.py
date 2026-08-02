#!/usr/bin/env python3
"""Create Wedding_Guest_List.xlsx with Summary + event sheets + Master list."""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# Mehndi guests from user's existing sheet (Family Members = headcount)
MEHNDI_GUESTS = [
    (1, "shree ganeshay namah", 0),
    (2, "shree bherulalji", 4),
    (3, "kamlesh shah", 6),
    (4, "dinesh jagetiya", 4),
    (5, "lalit jagetiya", 4),
    (6, "shree sundarlalji", 4),
    (7, "anil jagetiya", 2),
    (8, "rajesh jagetiya", 2),
    (9, "ganpat anand", 4),
    (10, "pintu anand", 4),
    (11, "manoj anand", 4),
    (12, "piyushbhai madhiwala", 3),
    (13, "sanjay ji bhopal", 9),
    (14, "nilesh bhatt", 5),
    (15, "nirav", 2),
    (16, "yogesh", 2),
    (17, "rakesh pathak", 1),
    (18, "shankarlalji", 0),
    (19, "shankarlalji mandowara family", 13),
    (20, "satyanarayanji mandowara", 7),
    (21, "radhika", 1),
    (22, "vishal jagetiya", 2),
    (23, "kantibhaishab", 3),
    (24, "jitindharmin", 5),
    (25, "upendrakumar family", 7),
    (26, "gunvantbhai", 5),
    (27, "jagdishbhai", 3),
    (28, "nimeshbhai kesariya", 2),
    (29, "darpan patel rakesh shah", 4),
    (30, "sanjay patel", 1),
    (31, "sanjay kothari", 3),
]

EVENTS = [
    ("Mehndi", "Mehndi (Mahendi)"),
    ("Haldi", "Haldi"),
    ("Paherwani", "Paherwani"),
    ("Ganpati Stapna", "Ganpati Stapna"),
    ("Sangeet", "Sangeet"),
    ("Wedding", "Wedding"),
]

# Sheet tab names (Excel-safe short names)
EVENT_SHEETS = [
    "Mehndi",
    "Haldi",
    "Paherwani",
    "Ganpati Stapna",
    "Sangeet",
    "Wedding",
]

HEADER_FILL = PatternFill("solid", fgColor="8B4557")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
TITLE_FONT = Font(bold=True, name="Georgia", size=20, color="5C2E3A")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="5C2E3A")
TOTAL_FILL = PatternFill("solid", fgColor="F3E6EA")
TOTAL_FONT = Font(bold=True, name="Calibri", size=12, color="5C2E3A")
LINK_FONT = Font(name="Calibri", size=12, color="8B4557", underline="single", bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D4C4C8"),
    right=Side(style="thin", color="D4C4C8"),
    top=Side(style="thin", color="D4C4C8"),
    bottom=Side(style="thin", color="D4C4C8"),
)
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
ALT_ROW = PatternFill("solid", fgColor="FAF6F7")
MAX_ROWS = 200  # editable rows for new guests


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_yes_no_validation(ws, start_row, end_row, start_col, end_col):
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.error = "Please select Yes or No"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        dv.add(f"{letter}{start_row}:{letter}{end_row}")


def build_workbook():
    wb = Workbook()

    # ----- Summary sheet -----
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Wedding Guest List"
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:D1")

    summary["A2"] = (
        "Click an event name below to open that sheet. "
        "Add or edit guests on the Master sheet (mark Yes for each event they attend), "
        "or add directly on an event sheet. Share this file via Google Sheets so family can edit together."
    )
    summary["A2"].font = SUBTITLE_FONT
    summary["A2"].alignment = Alignment(wrap_text=True)
    summary.merge_cells("A2:D2")
    summary.row_dimensions[2].height = 48

    summary["A4"] = "Event"
    summary["B4"] = "Open Sheet"
    summary["C4"] = "Guest Parties"
    summary["D4"] = "Total People"
    style_header_row(summary, 4, 4)

    # Event rows with hyperlinks + COUNTIF/SUMIF from Master
    # Master columns: A=S.No B=Guest Name C=Family Members D=Mehndi E=Haldi F=Paherwani G=Ganpati Stapna H=Sangeet I=Wedding J=Notes
    event_cols = {
        "Mehndi": "D",
        "Haldi": "E",
        "Paherwani": "F",
        "Ganpati Stapna": "G",
        "Sangeet": "H",
        "Wedding": "I",
    }

    for i, sheet_name in enumerate(EVENT_SHEETS):
        row = 5 + i
        col_letter = event_cols[sheet_name]
        summary.cell(row=row, column=1, value=sheet_name).font = NORMAL_FONT
        link_cell = summary.cell(row=row, column=2, value=f"→ Go to {sheet_name}")
        link_cell.font = LINK_FONT
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        # Count of Yes in that event column on Master
        summary.cell(
            row=row,
            column=3,
            value=f"=COUNTIF(Master!{col_letter}$2:{col_letter}${MAX_ROWS+1},\"Yes\")",
        )
        # Sum family members where that event is Yes
        summary.cell(
            row=row,
            column=4,
            value=f"=SUMIF(Master!{col_letter}$2:{col_letter}${MAX_ROWS+1},\"Yes\",Master!$C$2:$C${MAX_ROWS+1})",
        )
        for c in range(1, 5):
            summary.cell(row=row, column=c).border = THIN
            summary.cell(row=row, column=c).alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                summary.cell(row=row, column=c).fill = ALT_ROW

    summary["A12"] = "Grand Total (unique parties on Master)"
    summary["A12"].font = TOTAL_FONT
    summary["C12"] = f'=COUNTA(Master!B2:B{MAX_ROWS+1})'
    summary["D12"] = f'=SUM(Master!C2:C{MAX_ROWS+1})'
    for c in range(1, 5):
        summary.cell(row=12, column=c).fill = TOTAL_FILL
        summary.cell(row=12, column=c).font = TOTAL_FONT
        summary.cell(row=12, column=c).border = THIN

    summary["A14"] = "Quick instructions"
    summary["A14"].font = Font(bold=True, name="Georgia", size=14, color="5C2E3A")
    instructions = [
        "1. Prefer editing the Master sheet — one row per guest party.",
        "2. Set Family Members to the number of people in that party.",
        "3. Put Yes under every event they will attend (dropdowns provided).",
        "4. Each event sheet mirrors Master for that event and also lets you add rows.",
        "5. Upload to Google Sheets → Share → Anyone with the link → Editor.",
        "6. Mehndi guests from your existing list are already filled in.",
    ]
    for i, text in enumerate(instructions):
        summary.cell(row=15 + i, column=1, value=text).font = SUBTITLE_FONT
        summary.merge_cells(start_row=15 + i, start_column=1, end_row=15 + i, end_column=4)

    set_col_widths(summary, [36, 28, 16, 14])
    summary.row_dimensions[1].height = 28
    summary.freeze_panes = "A5"

    # ----- Master sheet -----
    master = wb.create_sheet("Master", 1)
    master["A1"] = "S.No"
    master["B1"] = "Guest Name"
    master["C1"] = "Family Members"
    master["D1"] = "Mehndi"
    master["E1"] = "Haldi"
    master["F1"] = "Paherwani"
    master["G1"] = "Ganpati Stapna"
    master["H1"] = "Sangeet"
    master["I1"] = "Wedding"
    master["J1"] = "Notes"
    style_header_row(master, 1, 10)

    for sno, name, members in MEHNDI_GUESTS:
        r = sno + 1  # data starts row 2
        master.cell(row=r, column=1, value=sno).font = NORMAL_FONT
        master.cell(row=r, column=2, value=name).font = NORMAL_FONT
        master.cell(row=r, column=3, value=members).font = NORMAL_FONT
        master.cell(row=r, column=4, value="Yes").font = NORMAL_FONT
        for c in range(5, 10):
            master.cell(row=r, column=c, value="No").font = NORMAL_FONT
        master.cell(row=r, column=10, value="").font = NORMAL_FONT
        for c in range(1, 11):
            master.cell(row=r, column=c).border = THIN
            if sno % 2 == 0:
                master.cell(row=r, column=c).fill = ALT_ROW

    # Pre-format empty rows for adding more guests
    for r in range(len(MEHNDI_GUESTS) + 2, MAX_ROWS + 2):
        master.cell(row=r, column=1, value=r - 1)  # auto S.No suggestion
        for c in range(1, 11):
            master.cell(row=r, column=c).border = THIN
            master.cell(row=r, column=c).font = NORMAL_FONT

    # Total row at bottom area
    total_row = MAX_ROWS + 2
    master.cell(row=total_row, column=2, value="TOTAL (all parties)").font = TOTAL_FONT
    master.cell(row=total_row, column=3, value=f"=SUM(C2:C{MAX_ROWS+1})").font = TOTAL_FONT
    for c in range(1, 11):
        master.cell(row=total_row, column=c).fill = TOTAL_FILL
        master.cell(row=total_row, column=c).border = THIN

    add_yes_no_validation(master, 2, MAX_ROWS + 1, 4, 9)
    set_col_widths(master, [8, 36, 16, 10, 10, 12, 16, 10, 10, 24])
    master.freeze_panes = "A2"
    master.auto_filter.ref = f"A1:J{MAX_ROWS+1}"

    # Highlight Yes cells in event columns
    for col in range(4, 10):
        letter = get_column_letter(col)
        master.conditional_formatting.add(
            f"{letter}2:{letter}{MAX_ROWS+1}",
            FormulaRule(formula=[f'{letter}2="Yes"'], fill=YES_FILL),
        )

    # ----- Per-event sheets -----
    # Each sheet shows guests for that event from Master via formulas,
    # plus a local "Quick Add" area for convenience. For Google Sheets compatibility,
    # we populate Mehndi with actual values and use FILTER-friendly layout:
    # Event sheets have their own editable list; Master is source of truth for multi-event.
    #
    # Practical approach for both Excel & Google Sheets:
    # - Event sheet columns: S.No | Guest Name | Family Members | Total People (formula)
    # - Pre-fill Mehndi from data
    # - Other events start empty with room to add
    # - Side panel shows running total
    # - Link back to Summary and Master

    for sheet_name in EVENT_SHEETS:
        ws = wb.create_sheet(sheet_name)
        event_col = event_cols[sheet_name]

        ws["A1"] = sheet_name
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:C1")

        ws["A2"] = "← Back to Summary"
        ws["A2"].font = LINK_FONT
        ws["A2"].hyperlink = "#'Summary'!A1"

        ws["C2"] = "Open Master (multi-event)"
        ws["C2"].font = LINK_FONT
        ws["C2"].hyperlink = "#'Master'!A1"

        # Side total box
        ws["E1"] = "Total People"
        ws["E1"].font = HEADER_FONT
        ws["E1"].fill = HEADER_FILL
        ws["E1"].alignment = Alignment(horizontal="center")
        ws["F1"] = f"=SUMIF(Master!{event_col}$2:{event_col}${MAX_ROWS+1},\"Yes\",Master!$C$2:$C${MAX_ROWS+1})"
        ws["F1"].font = TOTAL_FONT
        ws["F1"].fill = TOTAL_FILL
        ws["F1"].alignment = Alignment(horizontal="center")

        ws["E2"] = "Guest Parties"
        ws["E2"].font = HEADER_FONT
        ws["E2"].fill = HEADER_FILL
        ws["E2"].alignment = Alignment(horizontal="center")
        ws["F2"] = f"=COUNTIF(Master!{event_col}$2:{event_col}${MAX_ROWS+1},\"Yes\")"
        ws["F2"].font = TOTAL_FONT
        ws["F2"].fill = TOTAL_FILL
        ws["F2"].alignment = Alignment(horizontal="center")

        ws["A3"] = (
            "List below is synced from Master (rows marked Yes for this event). "
            "To add/edit: go to Master, set Guest Name + Family Members, and put Yes in this event column. "
            "Or use Quick Add below — then also mark Yes on Master for multi-event tracking."
        )
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].alignment = Alignment(wrap_text=True)
        ws.merge_cells("A3:F3")
        ws.row_dimensions[3].height = 40

        # Synced list headers
        headers = ["S.No", "Guest Name", "Family Members"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=5, column=i, value=h)
        style_header_row(ws, 5, 3)

        # Pull matching rows from Master using INDEX/AGGREGATE pattern that works in Excel.
        # For Google Sheets, FILTER is better — we document both.
        # Simpler reliable approach: for Mehndi, write values; for all sheets use formulas
        # that scan Master for Yes in the event column.

        # Formula approach (Excel 365 / Google Sheets compatible with FILTER):
        # We'll write a note and use FILTER formulas for modern sheets.
        # Google Sheets: =FILTER(Master!A2:C201, Master!D2:D201="Yes")
        # Excel 365: same FILTER function exists.

        # Place FILTER formula in A6 spilling array (Excel 365 / Google Sheets)
        filter_formula = (
            f'=IFERROR(FILTER(Master!A2:C{MAX_ROWS+1},Master!{event_col}2:{event_col}{MAX_ROWS+1}="Yes"),'
            f'"No guests yet — add on Master and set {sheet_name}=Yes")'
        )
        ws["A6"] = filter_formula
        ws["A6"].font = NORMAL_FONT

        # Quick Add section for local notes (optional mirror)
        qa_start = 12
        ws.cell(row=qa_start, column=1, value="Quick Add (optional local notes — prefer Master)").font = Font(
            bold=True, name="Georgia", size=12, color="5C2E3A"
        )
        ws.merge_cells(start_row=qa_start, start_column=1, end_row=qa_start, end_column=3)

        for i, h in enumerate(["S.No", "Guest Name", "Family Members"], 1):
            ws.cell(row=qa_start + 1, column=i, value=h)
        style_header_row(ws, qa_start + 1, 3)

        # Pre-fill Quick Add for Mehndi with original names so user has editable copy too
        if sheet_name == "Mehndi":
            for sno, name, members in MEHNDI_GUESTS:
                r = qa_start + 1 + sno
                ws.cell(row=r, column=1, value=sno).font = NORMAL_FONT
                ws.cell(row=r, column=2, value=name).font = NORMAL_FONT
                ws.cell(row=r, column=3, value=members).font = NORMAL_FONT
                for c in range(1, 4):
                    ws.cell(row=r, column=c).border = THIN
                    if sno % 2 == 0:
                        ws.cell(row=r, column=c).fill = ALT_ROW
            # Local total for quick add
            last = qa_start + 1 + len(MEHNDI_GUESTS)
            ws.cell(row=last + 1, column=2, value="Quick Add Total").font = TOTAL_FONT
            ws.cell(row=last + 1, column=3, value=f"=SUM(C{qa_start+2}:C{last})").font = TOTAL_FONT
            for c in range(1, 4):
                ws.cell(row=last + 1, column=c).fill = TOTAL_FILL
                ws.cell(row=last + 1, column=c).border = THIN
            # Extra empty rows for more entries
            for r in range(last + 2, last + 52):
                ws.cell(row=r, column=1, value=r - (qa_start + 1))
                for c in range(1, 4):
                    ws.cell(row=r, column=c).border = THIN
        else:
            for i in range(1, 51):
                r = qa_start + 1 + i
                ws.cell(row=r, column=1, value=i)
                for c in range(1, 4):
                    ws.cell(row=r, column=c).border = THIN
                    ws.cell(row=r, column=c).font = NORMAL_FONT
            last = qa_start + 1 + 50
            ws.cell(row=last + 1, column=2, value="Quick Add Total").font = TOTAL_FONT
            ws.cell(row=last + 1, column=3, value=f"=SUM(C{qa_start+2}:C{last})").font = TOTAL_FONT
            for c in range(1, 4):
                ws.cell(row=last + 1, column=c).fill = TOTAL_FILL
                ws.cell(row=last + 1, column=c).border = THIN

        set_col_widths(ws, [8, 36, 16, 4, 14, 12])
        ws.freeze_panes = "A6"
        ws.row_dimensions[1].height = 28

    # ----- How to Share sheet -----
    share = wb.create_sheet("How to Share")
    share["A1"] = "Share this guest list with family (no special software needed)"
    share["A1"].font = TITLE_FONT
    share.merge_cells("A1:B1")

    steps = [
        ("", ""),
        ("Option A — Google Sheets (recommended)", ""),
        ("1", "Open Google Drive (drive.google.com) on your phone or computer and sign in."),
        ("2", "Upload this Excel file (Wedding_Guest_List.xlsx)."),
        ("3", "Right-click the file → Open with → Google Sheets."),
        ("4", "Click Share (top right) → General access → Anyone with the link → Editor."),
        ("5", "Copy the link and send it to family on WhatsApp."),
        ("6", "Everyone can add/edit guest names and family members; updates appear for all."),
        ("", ""),
        ("Option B — Microsoft Excel Online / OneDrive", ""),
        ("1", "Upload the file to OneDrive."),
        ("2", "Open in Excel for the web → Share → Anyone with the link can edit."),
        ("", ""),
        ("How to add a guest to multiple events", ""),
        ("1", "Go to the Master sheet."),
        ("2", "Add Guest Name and Family Members."),
        ("3", "Set Yes under every event they will attend (Mehndi, Haldi, etc.)."),
        ("4", "Summary and event sheets update automatically."),
        ("", ""),
        ("Mehndi note", ""),
        ("", "Your existing 31 Mehndi guests (116 total people) are already loaded."),
    ]
    for i, (num, text) in enumerate(steps, start=3):
        share.cell(row=i, column=1, value=num).font = Font(bold=True, color="8B4557")
        share.cell(row=i, column=2, value=text).font = SUBTITLE_FONT if not text.startswith("Option") and text != "How to add a guest to multiple events" and text != "Mehndi note" else Font(bold=True, name="Georgia", size=13, color="5C2E3A")

    set_col_widths(share, [6, 90])

    out = "/workspace/wedding-guest-list/Wedding_Guest_List.xlsx"
    wb.save(out)
    print(f"Saved {out}")
    print(f"Mehndi guests: {len(MEHNDI_GUESTS)}")
    print(f"Mehndi total people: {sum(m for _, _, m in MEHNDI_GUESTS)}")


if __name__ == "__main__":
    build_workbook()
