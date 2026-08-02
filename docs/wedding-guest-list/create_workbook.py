"""Create the collaborative wedding-functions guest-list workbook."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


OUTPUT_PATH = Path(__file__).with_name("Wedding_Functions_Guest_List.xlsx")
EVENTS = (
    "Mahendi",
    "Haldi",
    "Paherwani",
    "Ganpati Stapna",
    "Sangeet",
    "Wedding",
)
FIRST_DATA_ROW = 5
LAST_DATA_ROW = 304

# Transcribed from the user's Mahendi screenshot. The entered headcounts total 116.
MAHENDI_GUESTS = (
    ("shree ganeshay namah", 0),
    ("shree bherulalji", 4),
    ("kamlesh shah", 6),
    ("dinesh jagetiya", 4),
    ("lalit jagetiya", 4),
    ("shree sundarlalji", 4),
    ("anil jagetiya", 2),
    ("rajesh jagetiya", 2),
    ("ganpat anand", 4),
    ("pintu anand", 4),
    ("manoj anand", 4),
    ("piyushbhai madhiwala", 3),
    ("sanjay ji bhopal", 9),
    ("nilesh bhatt", 5),
    ("nirav", 2),
    ("yogesh", 2),
    ("rakesh pathak", 1),
    ("shankarlalji", 0),
    ("shankarlalji mandowara family", 13),
    ("satyanarayaji mandowara", 7),
    ("radhika", 1),
    ("vishal jagetiya", 2),
    ("kantibhaishab", 3),
    ("jitindharmin", 5),
    ("upendrakumar family", 7),
    ("gunvantbhai", 5),
    ("jagdishbhai", 3),
    ("nimeshbhai kesariya", 2),
    ("darpan patel rakesh shah", 4),
    ("sanjay patel", 1),
    ("sanjay kothari", 3),
)

NAVY = "17365D"
TEAL = "0F6B5D"
PALE_TEAL = "DDEBF7"
PALE_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GRAY = "666666"
RED = "C00000"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def apply_title_style(cell, size=18):
    cell.font = Font(name="Aptos Display", size=size, bold=True, color=NAVY)
    cell.alignment = Alignment(vertical="center")


def apply_header_style(cell):
    cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(bottom=THIN_GRAY)


def set_column_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def add_summary_sheet(workbook):
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.sheet_properties.tabColor = TEAL
    sheet.freeze_panes = "A8"
    sheet.sheet_view.zoomScale = 95

    sheet["A1"] = "Wedding Functions Guest List"
    apply_title_style(sheet["A1"], 20)
    sheet["A2"] = (
        "Choose a function below. Every guest list is editable and updates its "
        "guest count automatically."
    )
    sheet["A2"].font = Font(name="Aptos", size=11, color=GRAY)
    sheet["A3"] = (
        "Tip: choose existing names from the dropdown, or type a new name. "
        "Use Guest Directory to maintain shared names and default headcounts."
    )
    sheet["A3"].font = Font(name="Aptos", size=10, italic=True, color=GRAY)

    sheet["A5"] = "Share with family"
    sheet["A5"].font = Font(name="Aptos", size=12, bold=True, color=NAVY)
    sheet["A6"] = (
        "Upload this file to OneDrive → open in Excel Online → Share → Link "
        "settings → Anyone with the link → Can edit → Apply → Copy link."
    )
    sheet["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A6"].fill = PatternFill("solid", fgColor=PALE_YELLOW)
    sheet["A6"].border = Border(
        left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
    )

    headers = ("Function", "Open list", "Guest entries", "Total guests")
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=8, column=column, value=value)
        apply_header_style(cell)

    for index, event in enumerate(EVENTS, start=9):
        safe_event = event.replace("'", "''")
        sheet.cell(row=index, column=1, value=event)
        link = sheet.cell(
            row=index,
            column=2,
            value=f'=HYPERLINK("#\'{safe_event}\'!A1","Open guest list")',
        )
        link.font = Font(name="Aptos", color="0563C1", underline="single")
        sheet.cell(
            row=index,
            column=3,
            value=f'=COUNTIF(\'{safe_event}\'!$B$5:$B$304,"?*")',
        )
        sheet.cell(
            row=index,
            column=4,
            value=f"=SUM('{safe_event}'!$C$5:$C$304)",
        )
        for column in range(1, 5):
            sheet.cell(row=index, column=column).border = Border(bottom=THIN_GRAY)

    total_row = 9 + len(EVENTS)
    sheet.cell(row=total_row, column=1, value="All functions total")
    sheet.cell(row=total_row, column=1).font = Font(name="Aptos", bold=True)
    sheet.cell(row=total_row, column=3, value=f"=SUM(C9:C{total_row - 1})")
    sheet.cell(row=total_row, column=4, value=f"=SUM(D9:D{total_row - 1})")
    for column in range(1, 5):
        sheet.cell(row=total_row, column=column).fill = PatternFill(
            "solid", fgColor=PALE_TEAL
        )
        sheet.cell(row=total_row, column=column).border = Border(
            top=THIN_GRAY, bottom=THIN_GRAY
        )

    directory_row = total_row + 2
    directory_link = sheet.cell(
        row=directory_row,
        column=1,
        value='=HYPERLINK("#\'Guest Directory\'!A1","Open Guest Directory")',
    )
    directory_link.font = Font(
        name="Aptos", size=11, bold=True, color="0563C1", underline="single"
    )

    sheet["A19"] = "Important"
    sheet["A19"].font = Font(name="Aptos", bold=True, color=RED)
    sheet["A20"] = (
        "An “Anyone can edit” link lets anyone who receives it change or delete "
        "entries. Keep a backup copy."
    )
    sheet["A20"].alignment = Alignment(wrap_text=True, vertical="top")

    set_column_widths(sheet, {"A": 26, "B": 22, "C": 18, "D": 18})
    sheet.row_dimensions[1].height = 29
    sheet.row_dimensions[6].height = 46
    sheet.row_dimensions[8].height = 24
    sheet.row_dimensions[20].height = 40
    sheet.auto_filter.ref = f"A8:D{total_row - 1}"
    sheet.print_title_rows = "1:8"


def add_guest_directory(workbook):
    sheet = workbook.create_sheet("Guest Directory")
    sheet.sheet_properties.tabColor = "5B9BD5"
    sheet.freeze_panes = "A5"
    sheet.sheet_view.zoomScale = 95

    sheet["A1"] = "Guest Directory"
    apply_title_style(sheet["A1"])
    sheet["A2"] = (
        "Edit shared names and default headcounts here. Event sheets use this "
        "list for their guest-name dropdowns and automatic headcount suggestions."
    )
    sheet["A2"].font = Font(name="Aptos", size=10, color=GRAY)
    sheet["A3"] = '=HYPERLINK("#Summary!A1","← Back to Summary")'
    sheet["A3"].font = Font(color="0563C1", underline="single")

    headers = ("Guest Name", "Default Family Members / Total People", "Notes")
    for column, value in enumerate(headers, start=1):
        apply_header_style(sheet.cell(row=4, column=column, value=value))

    for offset, (name, headcount) in enumerate(MAHENDI_GUESTS):
        row = FIRST_DATA_ROW + offset
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=headcount)

    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        sheet.cell(row=row, column=2).number_format = "0"
        for column in range(1, 4):
            sheet.cell(row=row, column=column).border = Border(bottom=THIN_GRAY)

    count_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="999",
        allow_blank=True,
    )
    count_validation.promptTitle = "Enter a headcount"
    count_validation.prompt = "Enter a whole number from 0 to 999."
    count_validation.errorTitle = "Whole numbers only"
    count_validation.error = "Use a whole number from 0 to 999."
    count_validation.showInputMessage = True
    count_validation.showErrorMessage = True
    sheet.add_data_validation(count_validation)
    count_validation.add(f"B{FIRST_DATA_ROW}:B{LAST_DATA_ROW}")

    table = Table(
        displayName="GuestDirectoryTable", ref=f"A4:C{LAST_DATA_ROW}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    set_column_widths(sheet, {"A": 38, "B": 38, "C": 32})
    sheet.row_dimensions[1].height = 27
    sheet.row_dimensions[4].height = 34
    sheet.print_title_rows = "1:4"


def add_event_sheet(workbook, event):
    sheet = workbook.create_sheet(event)
    sheet.sheet_properties.tabColor = "70AD47"
    sheet.freeze_panes = "B5"
    sheet.sheet_view.zoomScale = 90

    sheet["A1"] = f"{event} Guest List"
    apply_title_style(sheet["A1"])
    sheet["A2"] = (
        "Pick an existing name from the dropdown or type a new name. The "
        "headcount fills from Guest Directory when available and remains editable."
    )
    sheet["A2"].font = Font(name="Aptos", size=10, color=GRAY)
    sheet["A3"] = '=HYPERLINK("#Summary!A1","← Back to Summary")'
    sheet["A3"].font = Font(color="0563C1", underline="single")

    sheet["F2"] = "Guest entries"
    sheet["G2"] = f'=COUNTIF(B{FIRST_DATA_ROW}:B{LAST_DATA_ROW},"?*")'
    sheet["F3"] = "Total guests"
    sheet["G3"] = f"=SUM(C{FIRST_DATA_ROW}:C{LAST_DATA_ROW})"
    for coordinate in ("F2", "F3"):
        sheet[coordinate].font = Font(name="Aptos", bold=True, color=WHITE)
        sheet[coordinate].fill = PatternFill("solid", fgColor=TEAL)
    for coordinate in ("G2", "G3"):
        sheet[coordinate].font = Font(name="Aptos", size=14, bold=True, color=NAVY)
        sheet[coordinate].fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet[coordinate].alignment = Alignment(horizontal="center")
    sheet["F4"] = "Count each party exactly as you want it included in the total."
    sheet["F4"].font = Font(name="Aptos", size=9, italic=True, color=GRAY)
    sheet["F4"].alignment = Alignment(wrap_text=True)

    headers = (
        "S.No",
        "Guest Name",
        "Family Members / Total People",
        "Notes",
    )
    for column, value in enumerate(headers, start=1):
        apply_header_style(sheet.cell(row=4, column=column, value=value))

    imported = MAHENDI_GUESTS if event == "Mahendi" else ()
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        guest_index = row - FIRST_DATA_ROW
        sheet.cell(
            row=row,
            column=1,
            value=f'=IF(B{row}="","",ROW()-{FIRST_DATA_ROW - 1})',
        )
        if guest_index < len(imported):
            name, headcount = imported[guest_index]
            sheet.cell(row=row, column=2, value=name)
            sheet.cell(row=row, column=3, value=headcount)
        else:
            sheet.cell(
                row=row,
                column=3,
                value=(
                    f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},'
                    f"'Guest Directory'!$A$5:$B$304,2,FALSE),\"\"))"
                ),
            )
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        sheet.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        sheet.cell(row=row, column=3).number_format = "0"
        for column in range(1, 5):
            sheet.cell(row=row, column=column).border = Border(bottom=THIN_GRAY)

    guest_validation = DataValidation(
        type="list",
        formula1="=GuestDirectoryNames",
        allow_blank=True,
    )
    guest_validation.promptTitle = "Choose or type a guest"
    guest_validation.prompt = (
        "Choose a shared guest, or type a new name. Add new shared names to "
        "Guest Directory."
    )
    guest_validation.showInputMessage = True
    guest_validation.showErrorMessage = False
    sheet.add_data_validation(guest_validation)
    guest_validation.add(f"B{FIRST_DATA_ROW}:B{LAST_DATA_ROW}")

    count_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="999",
        allow_blank=True,
    )
    count_validation.promptTitle = "Enter a headcount"
    count_validation.prompt = "Enter a whole number from 0 to 999."
    count_validation.errorTitle = "Whole numbers only"
    count_validation.error = "Use a whole number from 0 to 999."
    count_validation.showInputMessage = True
    count_validation.showErrorMessage = True
    sheet.add_data_validation(count_validation)
    count_validation.add(f"C{FIRST_DATA_ROW}:C{LAST_DATA_ROW}")

    table_name = event.replace(" ", "") + "GuestTable"
    table = Table(displayName=table_name, ref=f"A4:D{LAST_DATA_ROW}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    sheet.conditional_formatting.add(
        f"C{FIRST_DATA_ROW}:C{LAST_DATA_ROW}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
        ),
    )

    set_column_widths(
        sheet,
        {"A": 9, "B": 38, "C": 34, "D": 32, "E": 3, "F": 26, "G": 15},
    )
    sheet.row_dimensions[1].height = 27
    sheet.row_dimensions[4].height = 34
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:G{LAST_DATA_ROW}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def verify_source_data():
    assert len(MAHENDI_GUESTS) == 31
    assert sum(headcount for _, headcount in MAHENDI_GUESTS) == 116
    assert len({name.casefold() for name, _ in MAHENDI_GUESTS}) == 31


def verify_workbook(path):
    workbook = load_workbook(path, data_only=False)
    expected_sheets = ["Summary", "Guest Directory", *EVENTS]
    assert workbook.sheetnames == expected_sheets
    assert workbook["Mahendi"]["B5"].value == "shree ganeshay namah"
    assert workbook["Mahendi"]["B35"].value == "sanjay kothari"
    assert sum(
        workbook["Mahendi"].cell(row=row, column=3).value
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + len(MAHENDI_GUESTS))
    ) == 116
    assert workbook["Mahendi"]["G3"].value == "=SUM(C5:C304)"
    assert workbook["Wedding"]["C5"].value.startswith("=IF(B5=")
    assert workbook["Summary"]["B9"].value.startswith("=HYPERLINK(")
    assert "GuestDirectoryNames" in workbook.defined_names
    for event in EVENTS:
        assert len(workbook[event].data_validations.dataValidation) == 2
        assert len(workbook[event].tables) == 1


def main():
    verify_source_data()

    workbook = Workbook()
    workbook.properties.creator = "Cursor"
    workbook.properties.title = "Wedding Functions Guest List"
    workbook.properties.subject = "Editable guest lists for six wedding functions"
    workbook.properties.description = (
        "Excel Online-compatible collaborative guest-list workbook."
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    add_summary_sheet(workbook)
    add_guest_directory(workbook)
    workbook.defined_names.add(
        DefinedName(
            "GuestDirectoryNames",
            attr_text=f"'Guest Directory'!$A${FIRST_DATA_ROW}:$A${LAST_DATA_ROW}",
        )
    )
    for event in EVENTS:
        add_event_sheet(workbook, event)

    workbook.active = 0
    workbook.save(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(
        f"Created {OUTPUT_PATH} with {len(EVENTS)} event sheets, "
        f"{len(MAHENDI_GUESTS)} imported Mahendi guests, and total headcount 116."
    )


if __name__ == "__main__":
    main()
