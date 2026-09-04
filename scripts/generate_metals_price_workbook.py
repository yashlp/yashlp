#!/usr/bin/env python3
"""Generate the Jagetiya Metals product price workbook from the built-in catalog.

Reads public/metals/js/catalog.js (via Node) so grades, shapes, subtypes, and
sizes stay in sync with the stock-search site.

Usage:
  python3 scripts/generate_metals_price_workbook.py
  python3 scripts/generate_metals_price_workbook.py --verify
  python3 scripts/generate_metals_price_workbook.py --verify-only

Grouping: one sheet per catalog grade name (`g`). Round, Square, Hex, Flat, and
Non-Ferrous for that grade share the sheet (MS Bright square/hex/flat together;
EN-8 round+square+flat together). MS, MS Bright, and MS Black stay separate.
EN-8 vs EN-8D vs EN-8D / C-45 stay separate because their `g` strings differ.

Daily adjustment: each grade sheet has a yellow cell at I3. Selling Price is
Base Price + $I$3, so typing 1 adds Rs 1 to every size of every shape on that
sheet without changing bases.

Add Size: green inbox (shape + SIZE / THICKNESS / WIDTH). New sizes append into
that shape's section on the same sheet via UNIQUE/VSTACK/FILTER. Duplicates
within the shape section (25 vs 25.0, 6x25 vs 6 x 25) are skipped.
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
CATALOG_JS = ROOT / "public" / "metals" / "js" / "catalog.js"
DEFAULT_OUTPUT = ROOT / "public" / "metals" / "Jagetiya_Metals_Price_List.xlsx"

NAVY = "0F2540"
GOLD = "C8960C"
CREAM = "F7F4EE"
WHITE = "FFFFFF"
YELLOW = "FFE566"
YELLOW_DARK = "F5C842"
PALE = "FFFDF8"
ROW_ALT = "FFF8E7"
LINK_BLUE = "1B3A5C"
MUTED = "5A6A7A"
GREEN = "1A5C32"
MINT = "E8F6EF"
MINT_INPUT = "F4FCF8"
SKY = "D6EAF8"

ADJ_CELL = "I3"
ADJ_ABS = "$I$3"
BODY_START = 6
PRODUCTS_SHEET = "Products"
EXAMPLE_BASES = {16: 72.00, 18: 71.50}  # first grade sheet only
EXTRA_SIZE_ROWS = 20
STAGING_FIRST_ROW = 7
STAGING_LAST_ROW = 26  # 20 optional extra size slots (L7:L26)
ADD_SHAPE_CELL = "M2"
ADD_SIZE_CELL = "M3"
ADD_TH_CELL = "M4"
ADD_W_CELL = "O4"
ADD_STATUS_CELL = "M5"
ADD_KEY_CELL = "P2"
SHAPE_CHOICES = ("Round Bar", "Square Bar", "Hex Bar", "Flat Bar", "Non-Ferrous")
SHAPE_ORDER = SHAPE_CHOICES
SHAPES_1D = ("Round Bar", "Hex Bar", "Non-Ferrous")
SECTION_TITLE = {
    "Round Bar": "ROUND BAR",
    "Square Bar": "SQUARE BAR",
    "Hex Bar": "HEX BAR",
    "Flat Bar": "FLAT BAR",
    "Non-Ferrous": "NON-FERROUS",
}
SECTION_HEADINGS = frozenset(SECTION_TITLE.values())

SHAPE_ABBR = {
    "Round Bar": "RB",
    "Square Bar": "SQ",
    "Hex Bar": "HX",
    "Flat Bar": "FL",
    "Non-Ferrous": "NF",
}
SHAPE_TAB = {
    "Round Bar": "1B3A5C",
    "Square Bar": "1A5C32",
    "Hex Bar": "6C3483",
    "Flat Bar": "C8960C",
    "Non-Ferrous": "117A65",
}
SUBTYPE_SHORT = (
    ("Centerless Grinding", "CL Grind"),
    ("Stainless Steel Rod", "SS Rod"),
    ("Rolled Imported", "Imp Rolled"),
    ("Imported / Forging Rod", "Imp Forg Rod"),
    ("Forging / Imported Rod", "Forg Imp Rod"),
    ("Turn Rod Imported", "Turn Rod Imp"),
    ("Rolled - Super Forge", "Rolled SF"),
    ("Rod, Hex, Square, Flat, Sheet", "Rod Hex Sq Fl"),
    ("HE30 / 6802 Grade", "HE30 6802"),
    ("Rod (Rolled)", "Rod Rolled"),
)

INVALID_SHEET_CHARS = re.compile(r'[:\\/*?\[\]]')
UNLOCKED = Protection(locked=False, hidden=False)

THIN = Border(
    left=Side(style="thin", color="D8D3CB"),
    right=Side(style="thin", color="D8D3CB"),
    top=Side(style="thin", color="D8D3CB"),
    bottom=Side(style="thin", color="D8D3CB"),
)
THICK_GOLD = Border(
    left=Side(style="medium", color=GOLD),
    right=Side(style="medium", color=GOLD),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)
THICK_GREEN = Border(
    left=Side(style="medium", color=GREEN),
    right=Side(style="medium", color=GREEN),
    top=Side(style="medium", color=GREEN),
    bottom=Side(style="medium", color=GREEN),
)
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
CREAM_FILL = PatternFill("solid", fgColor=CREAM)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
PALE_FILL = PatternFill("solid", fgColor=PALE)
ALT_FILL = PatternFill("solid", fgColor=ROW_ALT)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)
HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
MINT_FILL = PatternFill("solid", fgColor=MINT)
MINT_INPUT_FILL = PatternFill("solid", fgColor=MINT_INPUT)
SKY_FILL = PatternFill("solid", fgColor=SKY)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
COMPANY_FONT = Font(name="Calibri", size=11, bold=True, color=GOLD)
BODY_FONT = Font(name="Calibri", size=11, color=NAVY)
MUTED_FONT = Font(name="Calibri", size=10, color=MUTED)
GRADE_FONT = Font(name="Calibri", size=20, bold=True, color=NAVY)
COL_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
LINK_FONT = Font(name="Calibri", size=11, color=LINK_BLUE, underline="single", bold=True)
ADJ_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
ADJ_INPUT_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SELL_FONT = Font(name="Calibri", size=11, bold=True, color=GREEN)
EXAMPLE_FONT = Font(name="Calibri", size=10, italic=True, color="7D5A00")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def fmt_num(n) -> str:
    n = float(n)
    if abs(n - int(n)) < 1e-9:
        return str(int(n))
    return ("%g" % n)


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace("×", "x")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _compact_number(value):
    number = _as_number(value)
    if number is None:
        return None
    if abs(number - int(number)) < 1e-9:
        return int(number)
    return number


def normalize_existing_key(value):
    """Canonical form for duplicate checks: 25 == 25.0, '6x25' == '6 x 25'."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _compact_number(value)
    text = str(value).strip().replace(" ", "").replace("×", "x").replace("X", "x")
    if not text:
        return None
    if "x" in text:
        left, _, right = text.partition("x")
        a, b = _compact_number(left), _compact_number(right)
        if a is None or b is None:
            return text
        return "%sx%s" % (fmt_num(a), fmt_num(b))
    return _compact_number(text) if _as_number(text) is not None else text


def canonical_size_key(shape, size=None, thickness=None, width=None):
    """Size key produced by the Add Size inbox for the selected shape, or None."""
    shape = (shape or "").strip()
    size_n = _compact_number(size)
    th_n = _compact_number(thickness)
    w_n = _compact_number(width)
    if shape == "Flat Bar":
        if th_n is None or w_n is None:
            return None
        return "%sx%s" % (fmt_num(th_n), fmt_num(w_n))
    if shape == "Square Bar":
        if th_n is None:
            return None
        if w_n is None:
            return th_n
        return "%sx%s" % (fmt_num(th_n), fmt_num(w_n))
    if shape in SHAPES_1D:
        return size_n
    return None


def inbox_status(existing_keys, new_key):
    if new_key is None or new_key == "":
        return "Fill the fields for this shape"
    new_n = normalize_existing_key(new_key)
    for key in existing_keys:
        if normalize_existing_key(key) == new_n:
            return "Already added — skipped"
    return "Added"


def unique_size_list(catalog_keys, inbox_keys):
    """Catalog-first UNIQUE union used by extra rows (skip blanks and duplicates)."""
    out = []
    seen = set()
    for key in list(catalog_keys) + list(inbox_keys):
        norm = normalize_existing_key(key)
        if norm is None or norm == "":
            continue
        if norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def new_sizes_only(catalog_keys, inbox_keys):
    catalog_norm = {normalize_existing_key(k) for k in catalog_keys}
    catalog_norm.discard(None)
    catalog_norm.discard("")
    return [k for k in unique_size_list(catalog_keys, inbox_keys) if k not in catalog_norm]


def _unique_news_let(cat_ref: str) -> str:
    return (
        "LET(cat,%s,box,VSTACK($P$2,$L$7:$L$26),"
        "live,UNIQUE(FILTER(VSTACK(cat,box),VSTACK(cat,box)<>\"\")),"
        "FILTER(live,COUNTIF(cat,live)=0))"
    ) % cat_ref


def extra_size_formula(index: int, cat_ref: str, shape: str | None = None) -> str:
    """1-based INDEX into UNIQUE(VSTACK(catalog, inbox)) minus catalog (Excel 365 / Sheets).

    When `shape` is set, the inbox only feeds this section while M2 matches that shape.
    """
    let_expr = _unique_news_let(cat_ref)
    indexed = "IFERROR(INDEX(%s,%d),\"\")" % (let_expr, index)
    if shape:
        return '=IF($M$2<>"%s","",%s)' % (shape.replace('"', '""'), indexed)
    return "=" + indexed


def extra_flat_part_formula(index: int, cat_ref: str, part: str, shape: str | None = None) -> str:
    """Parse TxW key from the UNIQUE new-size list into thickness or width."""
    news = _unique_news_let(cat_ref)
    if part == "th":
        extract = "IFERROR(VALUE(LEFT(t,FIND(\"x\",t)-1)),\"\")"
    elif part == "w":
        extract = "IFERROR(VALUE(MID(t,FIND(\"x\",t)+1,32)),\"\")"
    else:
        extract = "k"
    body = (
        "LET(k,IFERROR(INDEX(%s,%d),\"\"),t,SUBSTITUTE(SUBSTITUTE(TRIM(k&\"\"),\" \",\"\"),\"×\",\"x\"),"
        "IF(OR(k=\"\",ISNUMBER(k)),\"\",%s))"
    ) % (news, index, extract)
    if shape:
        return '=IF($M$2<>"%s","",%s)' % (shape.replace('"', '""'), body)
    return "=" + body


def size_key_formula() -> str:
    n_m4 = "IF(ISNUMBER($M$4),$M$4,VALUE($M$4))"
    n_o4 = "IF(ISNUMBER($O$4),$O$4,VALUE($O$4))"
    n_m3 = "IF(ISNUMBER($M$3),$M$3,VALUE($M$3))"
    tw = "TEXT(%s,\"0.######\")&\"x\"&TEXT(%s,\"0.######\")" % (n_m4, n_o4)
    return (
        "=IFERROR(IF($M$2=\"Flat Bar\",IF(OR($M$4=\"\",$O$4=\"\"),\"\",%s),"
        "IF($M$2=\"Square Bar\",IF($M$4=\"\",\"\",IF($O$4=\"\",%s,%s)),"
        "IF(OR($M$2=\"Round Bar\",$M$2=\"Hex Bar\",$M$2=\"Non-Ferrous\"),"
        "IF($M$3=\"\",\"\",%s),\"\"))),\"\")"
    ) % (tw, n_m4, tw, n_m3)


def status_formula(shape_cat_refs) -> str:
    """Added vs skipped, checked only inside the selected shape's catalog range."""
    if isinstance(shape_cat_refs, str):
        shape_cat_refs = {"_": shape_cat_refs}
        inner = (
            'IF(OR(COUNTIF(%s,$P$2)>0,COUNTIF($L$7:$L$26,$P$2)>0),'
            '"Already added — skipped","Added")'
        ) % shape_cat_refs["_"]
        return '=IF($P$2="","Fill the fields for this shape",%s)' % inner
    inner = '"Pick a shape that exists on this sheet"'
    for shape, cat_ref in reversed(list(shape_cat_refs.items())):
        inner = (
            'IF($M$2="%s",IF(OR(COUNTIF(%s,$P$2)>0,COUNTIF($L$7:$L$26,$P$2)>0),'
            '"Already added — skipped","Added"),%s)'
        ) % (shape.replace('"', '""'), cat_ref, inner)
    return '=IF($P$2="","Fill the fields for this shape",%s)' % inner


def quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def load_catalog(catalog_js: Path) -> list[dict]:
    """Load builtin catalog entries via Node so JS stays the source of truth."""
    if not catalog_js.is_file():
        raise FileNotFoundError("Catalog not found: %s" % catalog_js)
    js = r"""
const c = require(%s);
const entries = [];
for (const shape of c.SL) {
  const list = c.BUILTIN_DB[shape] || [];
  for (const e of list) {
    const rec = {
      shape: shape,
      grade: e.g,
      subtype: e.s || "",
      make: e.m || "",
      sizes: Array.isArray(e.sz) ? e.sz.slice() : [],
      flat: e.flat ? e.flat : null,
      noteOnly: !!e.note
    };
    entries.push(rec);
  }
}
process.stdout.write(JSON.stringify(entries));
""" % json.dumps(str(catalog_js))
    proc = subprocess.run(
        ["node", "-e", js],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError("Failed to read catalog.js:\n%s" % proc.stderr)
    entries = json.loads(proc.stdout)
    if not entries:
        raise RuntimeError("Catalog parsed empty")
    return entries


def shorten_subtype(subtype: str) -> str:
    out = subtype
    for src, dst in SUBTYPE_SHORT:
        out = out.replace(src, dst)
    return out


def sanitize_sheet_name(raw: str) -> str:
    name = INVALID_SHEET_CHARS.sub("-", raw)
    name = re.sub(r"\s+", " ", name).strip()
    name = name.strip("'")
    if not name:
        name = "Grade"
    return name[:31]


def unique_grade_sheet_name(grade: str, used: set[str]) -> str:
    """Sheet name = sanitized grade (max 31). Suffix if two grades collide."""
    base = sanitize_sheet_name(grade)
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 50):
        suffix = "-%d" % i
        cand = sanitize_sheet_name(base[: 31 - len(suffix)] + suffix)
        if cand not in used:
            used.add(cand)
            return cand
    raise RuntimeError("Could not uniquify sheet name for %s" % grade)


def group_entries_by_grade(entries: list[dict]) -> list[dict]:
    """One family per catalog `g` string. Do not merge similar names."""
    order: list[str] = []
    by_grade: dict[str, list[dict]] = {}
    for e in entries:
        g = (e.get("grade") or "").strip()
        if not g:
            g = "Grade"
        if g not in by_grade:
            by_grade[g] = []
            order.append(g)
        by_grade[g].append(e)
    families = []
    for g in order:
        items = by_grade[g]
        present = []
        for e in items:
            if e["shape"] not in present:
                present.append(e["shape"])
        shapes = [s for s in SHAPE_ORDER if s in present]
        families.append({
            "grade": g,
            "entries": items,
            "shapes": shapes,
            "shapes_label": " · ".join(shapes),
            "size_count": sum(size_count(e) for e in items),
        })
    return families


def assign_family_sheets(families: list[dict]) -> list[dict]:
    used = {PRODUCTS_SHEET}
    for fam in families:
        fam["sheet"] = unique_grade_sheet_name(fam["grade"], used)
    return families


def classify(entry: dict) -> str:
    if entry.get("flat"):
        return "flat"
    if entry.get("noteOnly") and not entry.get("sizes"):
        return "note"
    return "round"


def size_count(entry: dict) -> int:
    kind = classify(entry)
    if kind == "flat":
        n = 0
        for widths in entry["flat"].values():
            n += len(widths)
        return n
    if kind == "note":
        return 0
    return len(entry.get("sizes") or [])


def flat_rows(entry: dict) -> list[tuple]:
    rows = []
    thicknesses = sorted((float(t) for t in entry["flat"].keys()), key=lambda x: x)
    for th in thicknesses:
        key = str(int(th)) if th == int(th) else str(th)
        if key not in entry["flat"]:
            key = str(th)
            if key not in entry["flat"]:
                # JS object keys are strings of the original number
                for k in entry["flat"]:
                    if abs(float(k) - th) < 1e-9:
                        key = k
                        break
        widths = sorted(entry["flat"][key], key=lambda w: float(w))
        for w in widths:
            label = "%sx%s" % (fmt_num(th), fmt_num(w))
            rows.append((th, w, label))
    return rows


def apply_cell(ws: Worksheet, row: int, col: int, value, *, font=None, fill=None,
               alignment=None, border=None, num_fmt=None, comment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.protection = UNLOCKED
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if num_fmt is not None:
        cell.number_format = num_fmt
    if comment is not None:
        cell.comment = comment
    return cell


def merge_fill(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, fill, font=None, alignment=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.protection = UNLOCKED
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
    return ws.cell(row=r1, column=c1)


def unlock_sheet(ws: Worksheet, max_row: int, max_col: int):
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.protection = UNLOCKED
    ws.protection.sheet = False
    ws.protection.enable = False
    ws.protection.autoFilter = True


def page_setup(ws: Worksheet, freeze: str = "A6"):
    ws.freeze_panes = freeze
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.55, bottom=0.5, header=0.25, footer=0.25)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddHeader.left.text = "Jagetiya Metals — Product Price List"
    ws.oddFooter.left.text = "GST 24AGIPS3207M1Z7  |  Fully editable — no password"
    ws.oddFooter.right.text = "Page &P of &N"


def add_adjustment_box(ws: Worksheet):
    """Yellow daily-adjustment box on the RIGHT of every grade sheet (I3)."""
    merge_fill(ws, 1, 8, 1, 10, GOLD_FILL, Font(name="Calibri", size=11, bold=True, color=NAVY), CENTER)
    ws.cell(row=1, column=8).value = "DAILY ADJUSTMENT (Rs/kg)"

    merge_fill(ws, 2, 8, 2, 10, YELLOW_FILL, Font(name="Calibri", size=9, color=NAVY), CENTER)
    ws.cell(row=2, column=8).value = (
        "Enter 1 to add Rs 1 to every size. Enter -2 to reduce Rs 2. Leave 0 for no change."
    )
    ws.row_dimensions[2].height = 32

    apply_cell(ws, 3, 8, "Today's change", font=ADJ_LABEL_FONT, fill=YELLOW_FILL, alignment=RIGHT)
    apply_cell(
        ws, 3, 9, 0,
        font=ADJ_INPUT_FONT,
        fill=YELLOW_FILL,
        alignment=CENTER,
        border=THICK_GOLD,
        num_fmt="0.00",
        comment=Comment(
            "Type today's market move here (example: 1 or -2).\n"
            "Every Selling Price on this sheet = Base Price + this cell.\n"
            "Base prices do not change. The delta does not compound.",
            "Jagetiya Metals",
        ),
    )
    apply_cell(ws, 3, 10, "Rs/kg", font=ADJ_LABEL_FONT, fill=YELLOW_FILL, alignment=LEFT)

    merge_fill(ws, 4, 8, 4, 10, YELLOW_FILL, Font(name="Calibri", size=9, italic=True, color=NAVY), CENTER)
    ws.cell(row=4, column=8).value = (
        "Selling Price = Base Price + this box. Change 1 to 3 and all sizes use +3 (not compounding)."
    )
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 28

    for r in range(1, 5):
        for c in range(8, 11):
            ws.cell(row=r, column=c).border = THICK_GOLD
            ws.cell(row=r, column=c).protection = UNLOCKED
    # re-apply inner I3 thick gold so it stays the obvious input
    ws["I3"].border = THICK_GOLD
    ws["I3"].fill = YELLOW_FILL
    ws["I3"].font = ADJ_INPUT_FONT
    ws["I3"].number_format = "0.00"


def add_size_box(ws: Worksheet, default_shape: str, status_val: str):
    """Green Add Size inbox to the right of Daily Adjustment (visible in frozen rows 1–5)."""
    title_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    hint_font = Font(name="Calibri", size=8, color=NAVY)
    status_font = Font(name="Calibri", size=10, bold=True, color=GREEN)

    merge_fill(ws, 1, 12, 1, 15, GREEN_FILL, title_font, CENTER)
    ws.cell(row=1, column=12).value = "ADD SIZE"
    ws.cell(row=1, column=12).comment = Comment(
        "Press Enter after filling dimensions; the size appears in that shape's section on THIS sheet if it is new.\n"
        "Round / Hex / Non-Ferrous: SIZE (mm) only.\n"
        "Square: THICKNESS (Side) and optional WIDTH. Blank width adds the Side; both add TxW.\n"
        "Flat: THICKNESS and WIDTH required (6 and 25 → 6x25).\n"
        "Duplicates are skipped within that shape section (25 = 25.0, 6x25 = 6 x 25).",
        "Jagetiya Metals",
    )

    apply_cell(ws, 2, 12, "Shape", font=ADJ_LABEL_FONT, fill=MINT_FILL, alignment=RIGHT)
    apply_cell(
        ws, 2, 13, default_shape,
        font=ADJ_LABEL_FONT, fill=MINT_INPUT_FILL, alignment=CENTER, border=THICK_GREEN,
        comment=Comment(
            "Defaults to the first shape on this sheet. Changing Shape adds the size into that "
            "shape's section on THIS same sheet (not another tab).",
            "Jagetiya Metals",
        ),
    )
    merge_fill(ws, 2, 14, 2, 15, MINT_FILL, hint_font, CENTER)
    ws.cell(row=2, column=14).value = "Adds into that shape section on this sheet."

    apply_cell(ws, 3, 12, "SIZE (mm)", font=ADJ_LABEL_FONT, fill=MINT_FILL, alignment=RIGHT)
    apply_cell(
        ws, 3, 13, None,
        font=ADJ_INPUT_FONT, fill=WHITE_FILL, alignment=CENTER, border=THICK_GREEN,
        comment=Comment("Round / Hex / Non-Ferrous: type SIZE (mm) and press Enter.", "Jagetiya Metals"),
    )
    merge_fill(ws, 3, 14, 3, 15, MINT_FILL, hint_font, LEFT)
    ws.cell(row=3, column=14).value = (
        '=IF(OR($M$2="Round Bar",$M$2="Hex Bar",$M$2="Non-Ferrous"),'
        '"1D: fill SIZE only, then press Enter.",'
        '"Leave SIZE blank — use Thickness and Width.")'
    )

    apply_cell(ws, 4, 12, "THICKNESS (mm)", font=ADJ_LABEL_FONT, fill=MINT_FILL, alignment=RIGHT)
    apply_cell(
        ws, 4, 13, None,
        font=ADJ_INPUT_FONT, fill=WHITE_FILL, alignment=CENTER, border=THICK_GREEN,
        comment=Comment(
            "Square: Thickness is the Side if Width is blank. Flat: required with Width.",
            "Jagetiya Metals",
        ),
    )
    apply_cell(ws, 4, 14, "WIDTH (mm)", font=ADJ_LABEL_FONT, fill=MINT_FILL, alignment=RIGHT)
    apply_cell(
        ws, 4, 15, None,
        font=ADJ_INPUT_FONT, fill=WHITE_FILL, alignment=CENTER, border=THICK_GREEN,
        comment=Comment(
            "Square: optional (blank = Side only; both = TxW). Flat: required, e.g. 6 and 25 → 6x25.",
            "Jagetiya Metals",
        ),
    )

    apply_cell(ws, 5, 12, "Status", font=ADJ_LABEL_FONT, fill=MINT_FILL, alignment=RIGHT)
    merge_fill(ws, 5, 13, 5, 15, MINT_FILL, status_font, CENTER)
    ws.cell(row=5, column=13).value = status_val

    apply_cell(
        ws, 1, 16, "Size key (auto)",
        font=Font(name="Calibri", size=8, bold=True, color=WHITE),
        fill=GREEN_FILL, alignment=CENTER,
    )
    apply_cell(
        ws, 2, 16, size_key_formula(),
        font=ADJ_LABEL_FONT, fill=MINT_INPUT_FILL, alignment=CENTER, border=THICK_GREEN,
    )
    merge_fill(ws, 3, 16, 5, 16, MINT_FILL, Font(name="Calibri", size=8, italic=True, color=NAVY), CENTER)
    ws.cell(row=3, column=16).value = (
        "Press Enter after filling dimensions; the size appears in the list if it is new."
    )

    for r in range(1, 6):
        for c in range(12, 17):
            cell = ws.cell(row=r, column=c)
            cell.protection = UNLOCKED
            if c < 16 and r in (1, 5):
                cell.border = THICK_GREEN
    for addr in (ADD_SHAPE_CELL, ADD_SIZE_CELL, ADD_TH_CELL, ADD_W_CELL, ADD_KEY_CELL):
        ws[addr].border = THICK_GREEN
        ws[addr].protection = UNLOCKED
    ws[ADD_SIZE_CELL].fill = WHITE_FILL
    ws[ADD_TH_CELL].fill = WHITE_FILL
    ws[ADD_W_CELL].fill = WHITE_FILL
    ws[ADD_SHAPE_CELL].fill = MINT_INPUT_FILL
    ws[ADD_KEY_CELL].fill = MINT_INPUT_FILL

    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(SHAPE_CHOICES),
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Shape",
        error="Choose Round Bar, Square Bar, Hex Bar, Flat Bar, or Non-Ferrous.",
        promptTitle="Shape",
        prompt="Select the shape; the size is added to that section on this sheet.",
        showInputMessage=True,
    )
    dv.add(ADD_SHAPE_CELL)
    ws.add_data_validation(dv)

    merge_fill(ws, 6, 12, 6, 15, GREEN_FILL, Font(name="Calibri", size=9, bold=True, color=WHITE), LEFT)
    ws.cell(row=6, column=12).value = "More sizes (optional) — type 25 or 6x25, one per row"
    for r in range(STAGING_FIRST_ROW, STAGING_LAST_ROW + 1):
        apply_cell(ws, r, 12, None, font=BODY_FONT, fill=MINT_INPUT_FILL, alignment=CENTER, border=THIN)
    apply_cell(
        ws, STAGING_FIRST_ROW, 13,
        "UNIQUE skips duplicates (25 = 25.0, 6x25 = 6 x 25)",
        font=MUTED_FONT, fill=MINT_FILL, alignment=LEFT,
    )


def add_products_sheet(wb: Workbook, families: list[dict]):
    ws = wb.active
    ws.title = PRODUCTS_SHEET
    ws.sheet_properties.tabColor = GOLD

    widths = {"A": 6, "B": 22, "C": 48, "D": 14, "E": 22, "F": 16, "G": 12, "H": 22, "I": 16, "J": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    merge_fill(ws, 1, 1, 1, 5, NAVY_FILL, TITLE_FONT, Alignment(horizontal="left", vertical="center", indent=1))
    ws.cell(row=1, column=1).value = "Jagetiya Metals — Product Price List"
    ws.row_dimensions[1].height = 32

    merge_fill(ws, 2, 1, 2, 5, NAVY_FILL, COMPANY_FONT, Alignment(horizontal="left", vertical="center", indent=1))
    ws.cell(row=2, column=1).value = (
        "Jagetiya Metals  ·  +91-9824012344  ·  Kamlesh@jkmetal.in  ·  "
        "502/1-A G.I.D.C., Makarpura, Vadodara  ·  GST 24AGIPS3207M1Z7"
    )
    ws.row_dimensions[2].height = 20

    merge_fill(ws, 3, 1, 3, 5, GOLD_FILL, Font(name="Calibri", size=10, bold=True, color=NAVY), LEFT)
    ws.cell(row=3, column=1).value = (
        "Fully editable workbook — no sheet protection, no locked cells, no password. "
        "One sheet per grade: Round, Square, Hex, and Flat for that grade live together."
    )
    ws.row_dimensions[3].height = 20

    instructions = [
        "How to use today's rates",
        "1. Open the grade sheet (click a name below, or the tabs at the bottom). Example: MS Bright has Square, Hex, and Flat on one tab.",
        "2. Put today's daily change in the yellow Daily Adjustment box on the RIGHT of that sheet (cell I3).",
        "3. Every size of every shape on that sheet updates: Selling Price = Base Price + Daily Adjustment.",
        "4. Type 1 to add Rs 1/kg to every size. Type -2 to reduce Rs 2/kg. Leave 0 for no change.",
        "5. Edit Base Price anytime — that is the stored rate. Daily Adjustment never overwrites it, and does not compound.",
        "6. Each grade has its own box so different grades can move independently. The yellow box on this sheet is only a copy-from hint.",
        "7. To add a size: use the green Add Size box. Pick the shape (Round / Square / Hex / Flat), fill SIZE or THICKNESS+WIDTH, press Enter.",
        "8. The size appears in that shape's section on THIS sheet if it is new. Duplicates within that section are skipped (25 = 25.0, 6x25 = 6 x 25).",
    ]
    for i, line in enumerate(instructions):
        r = 5 + i
        merge_fill(
            ws, r, 1, r, 5,
            CREAM_FILL if i else GOLD_FILL,
            Font(name="Calibri", size=11, bold=(i == 0), color=NAVY),
            LEFT,
        )
        ws.cell(row=r, column=1).value = line
        ws.row_dimensions[r].height = 18 if i else 22

    # Optional master daily adjustment — NOT linked to grade sheets
    merge_fill(ws, 5, 8, 5, 10, GOLD_FILL, Font(name="Calibri", size=11, bold=True, color=NAVY), CENTER)
    ws.cell(row=5, column=8).value = "MASTER DAILY ADJUSTMENT (optional)"
    merge_fill(ws, 6, 8, 7, 10, YELLOW_FILL, Font(name="Calibri", size=9, color=NAVY), CENTER)
    ws.cell(row=6, column=8).value = (
        "Copy this number into each grade sheet's yellow I3 box only if the same change applies. "
        "Grade sheets do NOT share this cell — each grade can move on its own."
    )
    apply_cell(ws, 8, 8, "Copy-from value", font=ADJ_LABEL_FONT, fill=YELLOW_FILL, alignment=RIGHT, border=THICK_GOLD)
    apply_cell(ws, 8, 9, 0, font=ADJ_INPUT_FONT, fill=YELLOW_FILL, alignment=CENTER, border=THICK_GOLD, num_fmt="0.00")
    apply_cell(ws, 8, 10, "Rs/kg", font=ADJ_LABEL_FONT, fill=YELLOW_FILL, alignment=LEFT, border=THICK_GOLD)
    merge_fill(ws, 9, 8, 11, 10, CREAM_FILL, MUTED_FONT, CENTER)
    ws.cell(row=9, column=8).value = "This cell is a notepad only. It does not drive grade selling prices."
    for r in range(5, 9):
        for c in range(8, 11):
            ws.cell(row=r, column=c).border = THICK_GOLD
            ws.cell(row=r, column=c).protection = UNLOCKED

    header_row = 15
    headers = ["#", "Grade", "Shapes on this sheet", "Size count", "Sheet name"]
    for c, h in enumerate(headers, 1):
        apply_cell(ws, header_row, c, h, font=COL_FONT, fill=HEADER_FILL, alignment=CENTER, border=THIN)
    ws.row_dimensions[header_row].height = 22
    ws.auto_filter.ref = "A%d:E%d" % (header_row, header_row + len(families))
    ws.freeze_panes = "A16"

    for i, fam in enumerate(families, 1):
        r = header_row + i
        fill = WHITE_FILL if i % 2 else ALT_FILL
        apply_cell(ws, r, 1, i, font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN)
        apply_cell(ws, r, 2, fam["grade"], font=Font(name="Calibri", size=11, bold=True, color=NAVY), fill=fill, alignment=LEFT, border=THIN)
        apply_cell(ws, r, 3, fam["shapes_label"], font=BODY_FONT, fill=fill, alignment=LEFT, border=THIN)
        apply_cell(ws, r, 4, fam["size_count"], font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN)
        link = apply_cell(ws, r, 5, fam["sheet"], font=LINK_FONT, fill=fill, alignment=LEFT, border=THIN)
        link.hyperlink = "#%s!A1" % quote_sheet(fam["sheet"])

    last = header_row + len(families)
    ws.auto_filter.ref = "A%d:E%d" % (header_row, last)
    note_r = last + 2
    merge_fill(ws, note_r, 1, note_r, 5, CREAM_FILL, MUTED_FONT, LEFT)
    ws.cell(row=note_r, column=1).value = (
        "Regenerate this file from the catalog:  python3 scripts/generate_metals_price_workbook.py"
    )
    page_setup(ws, freeze="A16")
    ws.print_title_rows = "1:15"
    unlock_sheet(ws, note_r + 2, 12)
    ws.protection.sheet = False


def _section_kind(entries: list[dict]) -> str:
    if any(classify(e) == "flat" for e in entries):
        return "flat"
    if entries and all(classify(e) == "note" for e in entries):
        return "note"
    return "round"


def add_grade_family_sheet(wb: Workbook, family: dict, is_first: bool):
    name = family["sheet"]
    ws = wb.create_sheet(title=name)
    shapes = family["shapes"]
    if len(shapes) == 1:
        ws.sheet_properties.tabColor = SHAPE_TAB.get(shapes[0], NAVY)
    else:
        ws.sheet_properties.tabColor = NAVY

    widths = {1: 22, 2: 28, 3: 14, 4: 16, 5: 16, 6: 18, 7: 16, 8: 22, 9: 14, 10: 14}
    extra_widths = {11: 3, 12: 22, 13: 16, 14: 16, 15: 14, 16: 18}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for col_idx, w in extra_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    last_table_col = 7
    merge_fill(ws, 1, 1, 1, last_table_col, NAVY_FILL, Font(name="Calibri", size=11, bold=True, color=GOLD), LEFT)
    ws.cell(row=1, column=1).value = "JAGETIYA METALS  ·  +91-9824012344  ·  Kamlesh@jkmetal.in  ·  GST 24AGIPS3207M1Z7"
    ws.row_dimensions[1].height = 20

    merge_fill(ws, 2, 1, 2, last_table_col, CREAM_FILL, GRADE_FONT, LEFT)
    ws.cell(row=2, column=1).value = family["grade"]
    ws.row_dimensions[2].height = 28

    merge_fill(ws, 3, 1, 3, last_table_col, CREAM_FILL, BODY_FONT, LEFT)
    ws.cell(row=3, column=1).value = "Shapes on this sheet: %s" % family["shapes_label"]

    back = apply_cell(
        ws, 4, 1,
        "← Back to Products index",
        font=LINK_FONT, fill=CREAM_FILL, alignment=LEFT,
    )
    back.hyperlink = "#%s!A1" % quote_sheet(PRODUCTS_SHEET)
    if is_first:
        merge_fill(ws, 4, 2, 4, last_table_col, PatternFill("solid", fgColor="FEF3CD"), EXAMPLE_FONT, LEFT)
        ws.cell(row=4, column=2).value = (
            "Example: sizes 16 and 18 have sample base prices. Put 1 in I3 — both selling prices rise by 1. "
            "Add a size with the green box; it lands in that shape's section on this sheet."
        )
    else:
        merge_fill(ws, 4, 2, 4, last_table_col, CREAM_FILL, MUTED_FONT, LEFT)
        ws.cell(row=4, column=2).value = (
            "Enter base prices in each shape section. Selling Price follows I3 for every size on this sheet. "
            "Add sizes with the green box — they append to the selected shape section."
        )

    merge_fill(ws, 5, 1, 5, last_table_col, CREAM_FILL, MUTED_FONT, LEFT)
    ws.cell(row=5, column=1).value = (
        "Daily Adjustment (I3) applies to all shapes below. Sub-type and make stay on the same sheet "
        "(example: EN-8D Rolled vs Rod)."
    )

    add_adjustment_box(ws)

    adj_abs = ADJ_ABS
    extra_note = "New size from Add Size — type a base price; selling uses $I$3"
    grouped = {s: [e for e in family["entries"] if e["shape"] == s] for s in shapes}
    shape_cat_refs: dict[str, str] = {}
    row = BODY_START
    example_used = False

    def write_price_row(r, kind, size_vals, subtype="", make="", example_base=None, note=""):
        fill = WHITE_FILL if r % 2 == 0 else ALT_FILL
        apply_cell(ws, r, 1, subtype, font=BODY_FONT, fill=fill, alignment=LEFT, border=THIN)
        apply_cell(ws, r, 2, make, font=MUTED_FONT, fill=fill, alignment=LEFT, border=THIN)
        if kind == "flat":
            apply_cell(ws, r, 3, size_vals.get("th"), font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN)
            apply_cell(ws, r, 4, size_vals.get("w"), font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN)
            apply_cell(
                ws, r, 5, size_vals.get("label"),
                font=Font(name="Calibri", size=11, bold=True, color=NAVY),
                fill=fill, alignment=CENTER, border=THIN,
            )
            base_col, adj_col, sell_col, notes_col = 6, 7, 8, None
        else:
            apply_cell(
                ws, r, 3, size_vals.get("size"),
                font=Font(name="Calibri", size=11, bold=True, color=NAVY),
                fill=fill, alignment=CENTER, border=THIN,
            )
            base_col, adj_col, sell_col, notes_col = 4, 5, 6, 7

        base_ref = "%s%d" % (get_column_letter(base_col), r)
        apply_cell(
            ws, r, base_col, example_base,
            font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN,
            num_fmt="0.00",
        )
        apply_cell(
            ws, r, adj_col, "=%s" % adj_abs,
            font=BODY_FONT, fill=fill, alignment=CENTER, border=THIN,
            num_fmt="0.00",
        )
        apply_cell(
            ws, r, sell_col, '=IF(%s="","",%s+%s)' % (base_ref, base_ref, adj_abs),
            font=SELL_FONT, fill=fill, alignment=CENTER, border=THIN,
            num_fmt="0.00",
        )
        if notes_col is not None:
            apply_cell(
                ws, r, notes_col, note,
                font=EXAMPLE_FONT if note else MUTED_FONT, fill=fill, alignment=LEFT, border=THIN,
            )

    for shape in shapes:
        entries = grouped[shape]
        kind = _section_kind(entries)
        title = SECTION_TITLE[shape]
        heading_cols = 8 if kind == "flat" else 7
        merge_fill(
            ws, row, 1, row, heading_cols, NAVY_FILL,
            Font(name="Calibri", size=13, bold=True, color=WHITE), LEFT,
        )
        ws.cell(row=row, column=1).value = title
        ws.row_dimensions[row].height = 22
        row += 1

        if kind == "flat":
            headers = [
                "Sub-type", "Make / notes", "Thickness (mm)", "Width (mm)", "Size",
                "Base Price (Rs/kg)", "Daily Adj", "Selling Price (Rs/kg)",
            ]
            size_col_letter = "E"
        else:
            headers = [
                "Sub-type", "Make / notes", "Size (mm)",
                "Base Price (Rs/kg)", "Daily Adj", "Selling Price (Rs/kg)", "Notes",
            ]
            size_col_letter = "C"

        for c, h in enumerate(headers, 1):
            apply_cell(ws, row, c, h, font=COL_FONT, fill=HEADER_FILL, alignment=CENTER, border=THIN)
        ws.row_dimensions[row].height = 22
        row += 1
        first_data = row

        for e in entries:
            subtype = e.get("subtype") or ""
            make = e.get("make") or ""
            if kind == "flat":
                for th, w, label in flat_rows(e):
                    write_price_row(row, "flat", {
                        "th": th if th != int(th) else int(th),
                        "w": w if float(w) != int(float(w)) else int(float(w)),
                        "label": label,
                    }, subtype=subtype, make=make)
                    row += 1
            elif kind == "note":
                note = make or "Available — add sizes with the green Add Size box (this Non-Ferrous section)"
                write_price_row(
                    row, "round", {"size": ""},
                    subtype=subtype,
                    make=note,
                    note="Catalog placeholder — use Add Size for new sizes",
                )
                row += 1
            else:
                for sz in e.get("sizes") or []:
                    num = float(sz)
                    display = int(num) if num == int(num) else num
                    example = None
                    note = ""
                    if is_first and (not example_used) and num in EXAMPLE_BASES:
                        example = EXAMPLE_BASES[num]
                        note = "Example — replace with live rate"
                    write_price_row(
                        row, "round", {"size": display},
                        subtype=subtype, make=make,
                        example_base=example, note=note,
                    )
                    row += 1
                    if example is not None and num == 18:
                        example_used = True

        if row == first_data:
            write_price_row(
                row, kind if kind == "flat" else "round",
                {"size": "", "th": "", "w": "", "label": ""},
                note="No catalog sizes — use Add Size",
            )
            row += 1

        last_cat = row - 1
        cat_ref = "$%s$%d:$%s$%d" % (size_col_letter, first_data, size_col_letter, last_cat)
        shape_cat_refs[shape] = cat_ref

        for i in range(1, EXTRA_SIZE_ROWS + 1):
            if kind == "flat":
                write_price_row(row, "flat", {
                    "th": extra_flat_part_formula(i, cat_ref, "th", shape),
                    "w": extra_flat_part_formula(i, cat_ref, "w", shape),
                    "label": extra_size_formula(i, cat_ref, shape),
                }, note=extra_note if i == 1 else "")
            else:
                write_price_row(
                    row, "round",
                    {"size": extra_size_formula(i, cat_ref, shape)},
                    note=extra_note if i == 1 else "",
                )
            row += 1
        row += 1  # spacer between sections

    default_shape = shapes[0] if shapes else "Round Bar"
    add_size_box(ws, default_shape, status_formula(shape_cat_refs))

    last_row = max(row, STAGING_LAST_ROW + 2)
    page_setup(ws, freeze="A6")
    ws.print_title_rows = "1:5"

    defn = DefinedName(
        name="DailyAdj",
        attr_text="%s!$I$3" % quote_sheet(name),
        localSheetId=None,
    )
    defn.localSheetId = wb.sheetnames.index(name)
    wb.defined_names.add(defn)

    unlock_sheet(ws, last_row, 16)
    ws.protection.sheet = False
    family["shape_cat_refs"] = shape_cat_refs
    family["last_row"] = last_row
    return ws


def build_workbook(entries: list[dict]) -> Workbook:
    families = assign_family_sheets(group_entries_by_grade(entries))

    wb = Workbook()
    wb.security.lockStructure = False
    wb.security.lockWindows = False
    wb.security.lockRevision = False
    wb.properties.title = "Jagetiya Metals — Product Price List"
    wb.properties.creator = "Jagetiya Metals"
    wb.properties.subject = "Editable steel price list by grade family (all shapes on one sheet)"
    wb.properties.description = (
        "One sheet per grade; Round/Square/Hex/Flat live together. "
        "Yellow Daily Adjustment (I3) applies to every size on that sheet. "
        "Green Add Size appends into the selected shape section. No protection, no password."
    )
    wb.properties.company = "Jagetiya Metals"

    add_products_sheet(wb, families)
    for i, fam in enumerate(families):
        add_grade_family_sheet(wb, fam, is_first=(i == 0))

    wb.security.lockStructure = False
    for ws in wb.worksheets:
        ws.protection.sheet = False
        ws.protection.enable = False
    return wb


def _verify_add_size_logic() -> list[str]:
    """Duplicate skip and shape-field rules (mirrors the UNIQUE inbox)."""
    errors = []
    cases = [
        ("Round Bar", 25, None, None, 25),
        ("Round Bar", 25.0, None, None, 25),
        ("Hex Bar", "25.0", None, None, 25),
        ("Non-Ferrous", 12, None, None, 12),
        ("Square Bar", None, 40, None, 40),
        ("Square Bar", None, 6, 25, "6x25"),
        ("Flat Bar", None, 6, 25, "6x25"),
        ("Flat Bar", None, 6, None, None),
        ("Round Bar", None, 16, None, None),
    ]
    for shape, size, th, w, expected in cases:
        got = canonical_size_key(shape, size, th, w)
        if got != expected:
            errors.append("canonical_size_key(%s) got %r want %r" % (shape, got, expected))

    catalog = [16, 18, 20, 25, 28]
    if inbox_status(catalog, 25) != "Already added — skipped":
        errors.append("duplicate 25 should be skipped")
    if inbox_status(catalog, 25.0) != "Already added — skipped":
        errors.append("duplicate 25.0 should be skipped")
    if inbox_status(catalog, "25.0") != "Already added — skipped":
        errors.append("duplicate text 25.0 should be skipped")
    if inbox_status(catalog, 30) != "Added":
        errors.append("new size 30 should be Added")
    if inbox_status(catalog, None) != "Fill the fields for this shape":
        errors.append("empty inbox should ask for fields")

    flats = ["6x25", "6x32", "10x40"]
    if inbox_status(flats, "6x25") != "Already added — skipped":
        errors.append("flat 6x25 duplicate")
    if inbox_status(flats, "6 x 25") != "Already added — skipped":
        errors.append("flat 6 x 25 should match 6x25")
    if inbox_status(flats, "6×25") != "Already added — skipped":
        errors.append("flat 6×25 should match 6x25")
    if inbox_status(flats, "8x25") != "Added":
        errors.append("new flat 8x25 should be Added")

    union = unique_size_list([16, 18, 25], [25.0, 30, "", 18])
    if union != [16, 18, 25, 30]:
        errors.append("unique_size_list order/dups got %r" % union)
    news = new_sizes_only([16, 18, 25], [25.0, 30, "18"])
    if news != [30]:
        errors.append("new_sizes_only should keep only 30, got %r" % news)
    return errors


def _is_section_heading_row(ws: Worksheet, r: int) -> bool:
    raw = ws.cell(r, 1).value
    title = raw.strip().upper() if isinstance(raw, str) else ""
    if title not in SECTION_HEADINGS:
        return False
    coord = ws.cell(r, 1).coordinate
    for mr in ws.merged_cells.ranges:
        if coord in mr and mr.min_row == r and mr.min_col == 1 and mr.max_col >= 5:
            return True
    return False


def iter_shape_sections(ws: Worksheet) -> list[dict]:
    """Parse stacked ROUND/SQUARE/HEX/FLAT/NON-FERROUS sections on a grade sheet."""
    sections = []
    max_r = ws.max_row or 6
    r = BODY_START
    heading_to_shape = {v: k for k, v in SECTION_TITLE.items()}
    while r <= max_r:
        if not _is_section_heading_row(ws, r):
            r += 1
            continue
        title = str(ws.cell(r, 1).value).strip().upper()
        shape = heading_to_shape[title]
        header_row = r + 1
        headers = [ws.cell(header_row, c).value for c in range(1, 9)]
        is_flat = headers[2] == "Thickness (mm)"
        if is_flat:
            size_col, base_col, adj_col, sell_col = 5, 6, 7, 8
        else:
            size_col, base_col, adj_col, sell_col = 3, 4, 5, 6
        data_start = header_row + 1
        end = data_start
        while end <= max_r:
            if _is_section_heading_row(ws, end):
                break
            end += 1
        sections.append({
            "shape": shape,
            "title": title,
            "header_row": header_row,
            "headers": headers,
            "is_flat": is_flat,
            "size_col": size_col,
            "base_col": base_col,
            "adj_col": adj_col,
            "sell_col": sell_col,
            "data_start": data_start,
            "data_end": end - 1,
        })
        r = end
    return sections


def _catalog_size_values(section: dict, ws: Worksheet) -> list:
    """Static catalog keys in a section (skip UNIQUE extra-row formulas)."""
    out = []
    for r in range(section["data_start"], section["data_end"] + 1):
        val = ws.cell(r, section["size_col"]).value
        if isinstance(val, str) and ("UNIQUE" in val or val.startswith("=")):
            continue
        if val in (None, ""):
            continue
        out.append(val)
    return out


def verify_workbook(path: Path, catalog_js: Path) -> dict:
    """Assert one sheet per grade family, stacked shapes, formulas, unlocked."""
    catalog = load_catalog(catalog_js)
    families = assign_family_sheets(group_entries_by_grade(catalog))
    unique_grades = []
    seen_g = set()
    for e in catalog:
        g = (e.get("grade") or "").strip()
        if g not in seen_g:
            seen_g.add(g)
            unique_grades.append(g)

    wb = load_workbook(path)
    errors = []
    report = {
        "path": str(path),
        "sheets": wb.sheetnames[:],
        "grade_sheets": len(wb.sheetnames) - 1,
        "catalog_entries": len(catalog),
        "unique_grades": len(families),
        "sample_formulas": {},
        "protected_sheets": [],
        "ok": True,
    }

    if PRODUCTS_SHEET not in wb.sheetnames:
        errors.append("Missing Products index sheet")
    if wb.sheetnames[0] != PRODUCTS_SHEET:
        errors.append("Products should be the first sheet, found %s" % wb.sheetnames[0])
    grade_n = len(wb.sheetnames) - 1
    if grade_n >= len(catalog):
        errors.append(
            "Grade sheet count %d should be fewer than catalog entries %d"
            % (grade_n, len(catalog))
        )
    if grade_n != len(families):
        errors.append(
            "Grade sheet count %d != unique grades %d"
            % (grade_n, len(families))
        )
    if len(families) != len(unique_grades):
        errors.append("Family count %d != unique grade names %d" % (len(families), len(unique_grades)))

    sec = wb.security
    if getattr(sec, "lockStructure", False):
        errors.append("Workbook structure is locked")
    if getattr(sec, "workbookPassword", None):
        errors.append("Workbook password is set")
    if getattr(sec, "revisionsPassword", None):
        errors.append("Revisions password is set")

    expected_names = {fam["sheet"] for fam in families}
    actual_grades = set(wb.sheetnames) - {PRODUCTS_SHEET}
    missing = expected_names - actual_grades
    extra = actual_grades - expected_names
    if missing:
        errors.append("Missing sheets: %s" % sorted(missing)[:8])
    if extra:
        errors.append("Unexpected sheets: %s" % sorted(extra)[:8])

    # Similar grade names must stay on separate sheets
    for keep_apart in (("EN-8", "EN-8D"), ("EN-8D", "EN-8D / C-45"), ("MS", "MS Bright"), ("MS Bright", "MS Black"), ("MS", "MS Black")):
        a, b = keep_apart
        sa = next((f["sheet"] for f in families if f["grade"] == a), None)
        sb = next((f["sheet"] for f in families if f["grade"] == b), None)
        if sa and sb and sa == sb:
            errors.append("Grades %r and %r were merged onto sheet %s" % (a, b, sa))

    adj_re = re.compile(r"=\$I\$3$", re.I)
    sell_re = re.compile(r'=IF\(([A-Z]+)(\d+)="","",\1\2\+\$I\$3\)$')
    unique_marks = ("UNIQUE", "VSTACK", "FILTER")
    shape_list = {s.strip() for s in SHAPE_CHOICES}

    first_grade = families[0]["sheet"] if families else None
    add_size_ok = 0
    extra_formula_ok = 0
    combined = {}

    for fam in families:
        name = fam["sheet"]
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        if ws.protection.sheet:
            report["protected_sheets"].append(name)
            errors.append("Sheet protection enabled on %s" % name)

        adj = ws[ADJ_CELL]
        if adj.value not in (0, 0.0):
            errors.append("%s %s default is %r, expected 0" % (name, ADJ_CELL, adj.value))
        if adj.number_format != "0.00":
            fmt = (adj.number_format or "").replace("\\", "")
            if "0.00" not in fmt:
                errors.append("%s %s number format is %r" % (name, ADJ_CELL, adj.number_format))

        title = ws.cell(1, 12).value
        if title != "ADD SIZE":
            errors.append("%s missing Add Size panel title, got %r" % (name, title))
        shape_val = ws[ADD_SHAPE_CELL].value
        if shape_val != fam["shapes"][0]:
            errors.append("%s Add Size shape default %r != first shape %r" % (name, shape_val, fam["shapes"][0]))
        if ws.cell(3, 12).value != "SIZE (mm)":
            errors.append("%s missing SIZE (mm) field" % name)
        if ws.cell(4, 12).value != "THICKNESS (mm)":
            errors.append("%s missing THICKNESS field" % name)
        if ws.cell(4, 14).value != "WIDTH (mm)":
            errors.append("%s missing WIDTH field" % name)
        status_val = ws[ADD_STATUS_CELL].value
        if not isinstance(status_val, str) or "Already added — skipped" not in status_val or "Added" not in status_val:
            errors.append("%s status formula missing Added/skipped text: %r" % (name, status_val))
        for present_shape in fam["shapes"]:
            if present_shape not in str(status_val):
                errors.append("%s status formula missing shape %r" % (name, present_shape))
                break
        key_val = ws[ADD_KEY_CELL].value
        if not isinstance(key_val, str) or "$M$2" not in key_val:
            errors.append("%s size-key formula missing shape switch: %r" % (name, key_val))
        else:
            if "Round Bar" not in key_val or "Flat Bar" not in key_val or "Square Bar" not in key_val:
                errors.append("%s size-key formula missing shape branches" % name)
            if "$M$3" not in key_val:
                errors.append("%s size-key formula missing SIZE input $M$3" % name)
            if "$M$4" not in key_val or "$O$4" not in key_val:
                errors.append("%s size-key formula missing THICKNESS/WIDTH cells" % name)
        dv_ok = False
        for dv in ws.data_validations.dataValidation:
            ranges = str(dv.sqref)
            if ADD_SHAPE_CELL in ranges and dv.type == "list":
                formula = dv.formula1 or ""
                if all(s in formula for s in shape_list):
                    dv_ok = True
        if not dv_ok:
            errors.append("%s missing shape dropdown on %s" % (name, ADD_SHAPE_CELL))
        else:
            add_size_ok += 1

        grade_cell = ws.cell(2, 1).value
        if grade_cell != fam["grade"]:
            errors.append("%s header grade %r != %r" % (name, grade_cell, fam["grade"]))
        shapes_line = str(ws.cell(3, 1).value or "")
        for s in fam["shapes"]:
            if s not in shapes_line:
                errors.append("%s header missing shape %r in %r" % (name, s, shapes_line))

        sections = iter_shape_sections(ws)
        found_shapes = [sec["shape"] for sec in sections]
        if found_shapes != fam["shapes"]:
            errors.append("%s sections %r != expected %r" % (name, found_shapes, fam["shapes"]))

        first_sell = None
        first_adj = None
        first_base = None
        extra_sections = 0
        section_sizes = {}

        for sec in sections:
            if sec["is_flat"]:
                if sec["headers"][2] != "Thickness (mm)" or sec["headers"][3] != "Width (mm)":
                    errors.append("%s %s missing thickness/width columns: %s" % (name, sec["title"], sec["headers"][:5]))
            else:
                if sec["headers"][2] != "Size (mm)":
                    errors.append("%s %s missing Size (mm) column: %s" % (name, sec["title"], sec["headers"][:4]))
            if sec["headers"][0] != "Sub-type":
                errors.append("%s %s missing Sub-type column" % (name, sec["title"]))

            catalog_keys = []
            extra_checked = 0
            seen_pairs = set()
            for r in range(sec["data_start"], sec["data_end"] + 1):
                size_val = ws.cell(r, sec["size_col"]).value
                adjf = ws.cell(r, sec["adj_col"]).value
                sell = ws.cell(r, sec["sell_col"]).value
                base = ws.cell(r, sec["base_col"]).value
                subtype = ws.cell(r, 1).value
                is_formula = isinstance(size_val, str) and (size_val.startswith("=") or "UNIQUE" in size_val)

                if size_val in (None, "") and not is_formula:
                    # spacer / empty note row: still require adj/sell if adj looks like a formula
                    if isinstance(adjf, str) and adjf.strip().startswith("="):
                        if not adj_re.match(adjf.strip()):
                            errors.append("%s %s row %d Daily Adj formula %r (want =$I$3)" % (name, sec["title"], r, adjf))
                            break
                        if not isinstance(sell, str) or not sell_re.match(sell.strip()):
                            errors.append("%s %s row %d Selling formula %r" % (name, sec["title"], r, sell))
                            break
                    continue

                if not isinstance(adjf, str) or not adj_re.match(str(adjf).strip()):
                    errors.append("%s %s row %d Daily Adj formula %r (want =$I$3)" % (name, sec["title"], r, adjf))
                    break
                if not isinstance(sell, str) or not sell_re.match(sell.strip()):
                    errors.append("%s %s row %d Selling formula %r" % (name, sec["title"], r, sell))
                    break
                if ws.cell(r, sec["sell_col"]).protection.locked:
                    errors.append("%s %s row %d selling cell is locked" % (name, sec["title"], r))
                    break
                if "$I$3" not in sell or "I3" in sell.replace("$I$3", ""):
                    errors.append("%s %s row %d selling formula must use absolute $I$3: %r" % (name, sec["title"], r, sell))
                    break

                if first_sell is None and not is_formula:
                    first_sell = sell
                    first_adj = adjf
                    first_base = base

                if is_formula:
                    missing_m = [m for m in unique_marks if m not in size_val]
                    if missing_m:
                        errors.append("%s %s extra row %d missing %s in %r" % (name, sec["title"], r, missing_m, size_val[:80]))
                        break
                    if "$M$2" not in size_val:
                        errors.append("%s %s extra row %d UNIQUE formula must gate on selected shape $M$2" % (name, sec["title"], r))
                        break
                    extra_checked += 1
                else:
                    catalog_keys.append(size_val)
                    key = normalize_existing_key(size_val)
                    pair = (str(subtype or ""), key)
                    if pair in seen_pairs:
                        errors.append("%s %s duplicate size %r subtype %r" % (name, sec["title"], size_val, subtype))
                    seen_pairs.add(pair)

            if extra_checked >= EXTRA_SIZE_ROWS:
                extra_sections += 1
            elif extra_checked:
                errors.append("%s %s has %d UNIQUE extra rows, want %d" % (name, sec["title"], extra_checked, EXTRA_SIZE_ROWS))

            if sec["is_flat"]:
                # first extra formula row: thickness/width parse
                for r in range(sec["data_start"], sec["data_end"] + 1):
                    th_f = ws.cell(r, 3).value
                    if isinstance(th_f, str) and "UNIQUE" in th_f or (isinstance(th_f, str) and th_f.startswith("=") and "FIND" in th_f):
                        if "FIND" not in str(th_f) or "x" not in str(th_f):
                            errors.append("%s extra flat thickness formula missing TxW parse: %r" % (name, th_f))
                        w_f = ws.cell(r, 4).value
                        if not (isinstance(w_f, str) and "FIND" in w_f):
                            errors.append("%s extra flat width formula missing TxW parse: %r" % (name, w_f))
                        break

            section_sizes[sec["shape"]] = catalog_keys

        if extra_sections == len(sections) and sections:
            extra_formula_ok += 1

        combined[fam["grade"]] = {
            "sheet": name,
            "shapes": found_shapes,
            "sizes": {sh: [normalize_existing_key(k) for k in keys] for sh, keys in section_sizes.items()},
        }

        if name == first_grade:
            report["sample_formulas"] = {
                "sheet": name,
                "I3": adj.value,
                "daily_adj": first_adj,
                "selling_price": first_sell,
                "example_base_16mm": first_base,
                "add_size_shape": shape_val,
                "add_size_status": status_val,
                "conceptual_if_adj_1": (
                    None if first_base in (None, "") else float(first_base) + 1
                ),
            }
            if first_base not in EXAMPLE_BASES.values():
                errors.append("First grade sheet should include example base prices, got %r" % first_base)

    report["add_size_panels"] = add_size_ok
    report["extra_unique_rows"] = extra_formula_ok
    report["combined"] = {g: {"sheet": v["sheet"], "shapes": v["shapes"]} for g, v in combined.items()}
    if add_size_ok != len(families):
        errors.append("Add Size panel present on %d / %d grade sheets" % (add_size_ok, len(families)))
    if extra_formula_ok != len(families):
        errors.append("UNIQUE extra rows present on %d / %d grade sheets" % (extra_formula_ok, len(families)))

    def _has_shapes(grade, needed):
        info = combined.get(grade) or {}
        have = set(info.get("shapes") or [])
        missing_s = [s for s in needed if s not in have]
        if missing_s:
            errors.append("%s sheet missing sections %s (have %s)" % (grade, missing_s, sorted(have)))
        return info

    ms = _has_shapes("MS Bright", ["Square Bar", "Hex Bar", "Flat Bar"])
    if ms:
        sq = set((combined.get("MS Bright") or {}).get("sizes", {}).get("Square Bar") or [])
        hx = set((combined.get("MS Bright") or {}).get("sizes", {}).get("Hex Bar") or [])
        fl = set((combined.get("MS Bright") or {}).get("sizes", {}).get("Flat Bar") or [])
        for expect in (8, 25.4, 63):
            if expect not in sq:
                errors.append("MS Bright Square missing size %s (have sample %s)" % (expect, list(sq)[:8]))
                break
        for expect in (12, 75):
            if expect not in hx:
                errors.append("MS Bright Hex missing size %s (have sample %s)" % (expect, list(hx)[:8]))
                break
        if "5x16" not in fl or "6x25" not in fl:
            errors.append("MS Bright Flat missing TxW sizes, have %s" % list(fl)[:8])

    _has_shapes("EN-8", ["Round Bar", "Flat Bar"])
    _has_shapes("WPS (D3)", ["Round Bar", "Square Bar", "Flat Bar"])

    logic_errors = _verify_add_size_logic()
    errors.extend(logic_errors)

    p = wb[PRODUCTS_SHEET]
    if p.protection.sheet:
        errors.append("Products sheet is protected")
        report["protected_sheets"].append(PRODUCTS_SHEET)
    products_header = 15
    hdr = [p.cell(products_header, c).value for c in range(1, 6)]
    if hdr != ["#", "Grade", "Shapes on this sheet", "Size count", "Sheet name"]:
        errors.append("Products headers %r" % hdr)
    found_links = 0
    for row in p.iter_rows(min_row=products_header + 1, max_row=products_header + len(families), min_col=5, max_col=5):
        cell = row[0]
        if cell.hyperlink:
            found_links += 1
    if found_links != len(families):
        errors.append("Products hyperlinks %d != %d grade sheets" % (found_links, len(families)))
    if p.cell(products_header + 1, 2).value != families[0]["grade"]:
        errors.append("Products first grade %r != %r" % (p.cell(products_header + 1, 2).value, families[0]["grade"]))

    sample = report["sample_formulas"]
    if sample.get("example_base_16mm") not in (None, ""):
        base = float(sample["example_base_16mm"])
        report["sample_formulas"]["conceptual_checks"] = {
            "base": base,
            "adj_1_selling": base + 1,
            "adj_3_selling": base + 3,
            "adj_minus2_selling": base - 2,
            "note": "Changing I3 from 1 to 3 yields base+3, not (base+1)+3",
        }

    report["errors"] = errors
    report["ok"] = not errors
    report["sheet_names"] = wb.sheetnames
    return report


def print_report(report: dict) -> None:
    print("Workbook:", report["path"])
    print("Catalog entries:", report["catalog_entries"])
    print("Grade sheets:", report["grade_sheets"])
    print("Total sheets (including Products):", len(report.get("sheet_names") or report["sheets"]))
    print("Sheet names:")
    for n in report.get("sheet_names") or report["sheets"]:
        print("  -", n)
    sf = report.get("sample_formulas") or {}
    if sf:
        print("Sample formulas (%s):" % sf.get("sheet"))
        for k, v in sf.items():
            if k != "sheet":
                print("  %s: %s" % (k, v))
    if report["errors"]:
        print("VERIFY FAILED:")
        for e in report["errors"]:
            print("  -", e)
    else:
        print("VERIFY OK — unprotected, one sheet per grade family, I3 daily adj, Add Size UNIQUE, Selling=Base+$I$3")
        if report.get("unique_grades"):
            print("Unique grades / sheets:", report["unique_grades"])
        if report.get("add_size_panels"):
            print("Add Size panels:", report["add_size_panels"])
        if report.get("extra_unique_rows"):
            print("UNIQUE extra-row sheets:", report["extra_unique_rows"])


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate Jagetiya Metals price workbook")
    parser.add_argument("--catalog", type=Path, default=CATALOG_JS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--verify", action="store_true", help="Generate then verify")
    parser.add_argument("--verify-only", action="store_true", help="Verify existing file only")
    parser.add_argument("--log", type=Path, default=None, help="Write JSON report to this path")
    args = parser.parse_args(argv)

    if not args.verify_only:
        entries = load_catalog(args.catalog)
        wb = build_workbook(entries)
        args.output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(args.output)
        print("Wrote %s (%d grade sheets from %d catalog entries)" % (
            args.output, len(wb.sheetnames) - 1, len(entries),
        ))

    if args.verify or args.verify_only:
        report = verify_workbook(args.output, args.catalog)
        print_report(report)
        if args.log:
            args.log.parent.mkdir(parents=True, exist_ok=True)
            args.log.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
            print("Log:", args.log)
        return 0 if report["ok"] else 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
